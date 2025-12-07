from openpyxl.styles import Font, Alignment
from requests_ntlm import HttpNtlmAuth
from datetime import datetime
import subprocess
import requests
import openpyxl
import argparse
import warnings
import urllib3
import base64
import sys
import re
from abc import ABC, abstractmethod

class VCSProvider(ABC):
    def __init__(self, args):
        self.args = args

    @abstractmethod
    def construct_api_url(self):
        pass

    @abstractmethod
    def get_auth_headers(self):
        pass

    @abstractmethod
    def fetch_commits(self, api_url, auth, headers):
        pass

    @abstractmethod
    def get_work_item_info(self, item_id, auth, headers):
        pass

    @abstractmethod
    def get_commit_ids_from_branches(self):
        pass

class TFSProvider(VCSProvider):
    def construct_api_url(self):
        return (
            f"https://msk-tfs-t.infotecs-nt/tfs/SrvNccCollection/{self.args.project}/_apis/git/repositories/{self.args.repository}/commits"
            f"?searchCriteria.itemVersion.version={self.args.branch}&searchCriteria.includeWorkItems=true&api-version=7.1-preview.1"
        )

    def get_auth_headers(self):
        auth = HttpNtlmAuth(self.args.username, self.args.password)
        headers = {'Content-Type': 'application/json-patch+json'}
        return auth, headers

    def fetch_commits(self, api_url, auth, headers):
        warnings.filterwarnings('ignore')
        try:
            response = requests.get(url=api_url, auth=auth, headers=headers, verify=False, timeout=10)
            response.raise_for_status()
            return response.json().get('value', [])
        except requests.RequestException as e:
            raise ValueError("Не удалось получить коммиты из TFS")

    def get_work_item_info(self, item_id, auth, headers):
        child_url = f"https://msk-tfs-t.infotecs-nt/tfs/SrvNccCollection/_apis/wit/workItems/{item_id}?$expand=relations&api-version=7.1"
        info_from_api = self.fetch_response(child_url, auth, headers)
        parent_id = info_from_api['fields'].get('System.Parent', '-')
        task_type = info_from_api['fields'].get('System.WorkItemType', '-')
        task_title = info_from_api['fields'].get('System.Title', '-')
        task_project = info_from_api['fields'].get('System.TeamProject', '-')
        task_url = f"https://msk-tfs-t.infotecs-nt/tfs/SrvNccCollection/{task_project}/_workitems/edit/{item_id}"
        return parent_id, item_id, task_type, task_title, task_url

    def fetch_response(self, url, auth, headers):
        response = requests.get(url=url, auth=auth, headers=headers, verify=False, timeout=10)
        response.raise_for_status()
        return response.json()

    def get_commit_ids_from_branches(self):
        repo_for_clone = f"https://msk-tfs-t.infotecs-nt/tfs/SrvNccCollection/{self.args.project}/_git/{self.args.repository}"
        base = base64.b64encode(bytes(f":{self.args.pat}", 'utf-8')).decode()
        cmd = f"git -c http.extraHeader=\"Authorization: Basic {base}\" clone {repo_for_clone}"
        git_clone = subprocess.run(cmd, shell=True)
        if git_clone.returncode != 0:
            print("Unable to locally clone TFS repo")
            sys.exit(1)

        cmd = f"cd {self.args.repository}; git cherry {self.args.previous_branch} origin/{self.args.branch}"
        git_cherry = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE)
        if git_cherry.returncode != 0:
            print(f"Unable to find differences between {self.args.previous_branch} and {self.args.branch}")
            sys.exit(1)

        cherry_commits = [line[2:] for line in reversed(git_cherry.stdout.decode().strip().split('\n'))]
        cmd = f"rm -rf ./{self.args.repository}"
        subprocess.run(cmd, shell=True)
        return cherry_commits

class GitHubProvider(VCSProvider):
    def construct_api_url(self):
        params = {}
        if self.args.search_by_date:
            params['since'] = self.args.start_date + 'T00:00:00Z'
            params['until'] = self.args.end_date + 'T23:59:59Z'
        if self.args.branch:
            params['sha'] = self.args.branch
        url = f"https://api.github.com/repos/{self.args.owner}/{self.args.repository}/commits"
        return url, params

    def get_auth_headers(self):
        headers = {
            'Authorization': f'token {self.args.github_token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        return None, headers  # No auth object, use headers directly

    def fetch_commits(self, api_url, auth, headers):
        try:
            params = self.construct_api_url()[1] if isinstance(self.construct_api_url(), tuple) else {}
            response = requests.get(url=api_url, headers=headers, params=params, timeout=10)
            response.raise_for_status()
            commits = response.json()
            # Normalize to TFS-like structure
            normalized_commits = []
            for c in commits:
                work_items = self.extract_work_items_from_message(c['commit']['message'])
                normalized_commits.append({
                    'commitId': c['sha'],
                    'author': {'name': c['commit']['author']['name'], 'date': c['commit']['author']['date']},
                    'comment': c['commit']['message'],
                    'workItems': [{'id': wi} for wi in work_items]
                })
            return normalized_commits
        except requests.RequestException as e:
            raise ValueError("Не удалось получить коммиты из GitHub")

    def extract_work_items_from_message(self, message):
        # Parse #123 or closes #123
        return re.findall(r'(?:#|closes #)(\d+)', message)

    def get_work_item_info(self, item_id, auth, headers):
        # Assume item_id is issue or PR ID
        # Try issue first
        issue_url = f"https://api.github.com/repos/{self.args.owner}/{self.args.repository}/issues/{item_id}"
        response = requests.get(issue_url, headers=headers)
        if response.status_code == 200:
            issue = response.json()
            parent_id = '-'  # No strict parent in GitHub, perhaps from linked issues
            task_type = 'Issue' if not issue.get('pull_request') else 'Pull Request'
            task_title = issue['title']
            task_url = issue['html_url']
            return parent_id, item_id, task_type, task_title, task_url

        # If not issue, assume PR
        pr_url = f"https://api.github.com/repos/{self.args.owner}/{self.args.repository}/pulls/{item_id}"
        response = requests.get(pr_url, headers=headers)
        if response.status_code == 200:
            pr = response.json()
            parent_id = '-'
            task_type = 'Pull Request'
            task_title = pr['title']
            task_url = pr['html_url']
            return parent_id, item_id, task_type, task_title, task_url

        return '-', item_id, '-', '-', '-'

    def get_commit_ids_from_branches(self):
        repo_for_clone = f"https://github.com/{self.args.owner}/{self.args.repository}.git"
        cmd = f"git clone {repo_for_clone}"
        git_clone = subprocess.run(cmd, shell=True)
        if git_clone.returncode != 0:
            print("Unable to locally clone GitHub repo")
            sys.exit(1)

        cmd = f"cd {self.args.repository}; git cherry {self.args.previous_branch} {self.args.branch}"
        git_cherry = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE)
        if git_cherry.returncode != 0:
            print(f"Unable to find differences between {self.args.previous_branch} and {self.args.branch}")
            sys.exit(1)

        cherry_commits = [line[2:] for line in reversed(git_cherry.stdout.decode().strip().split('\n'))]
        cmd = f"rm -rf ./{self.args.repository}"
        subprocess.run(cmd, shell=True)
        return cherry_commits

def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument('--platform', type=str, required=True, choices=['tfs', 'github'], help='Платформа: tfs или github')
    # Common args
    parser.add_argument('--repository', type=str, required=True, help='Репозиторий')
    parser.add_argument('--branch', type=str, required=False, help='Ветка')
    parser.add_argument('--search_by_date', type=str, required=True, help='Поиск по дате: True/False')
    parser.add_argument('--start_date', type=str, required=False, help='Начальная дата (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, required=False, help='Конечная дата (YYYY-MM-DD)')
    parser.add_argument('--previous_branch', type=str, required=False, help='Предыдущая ветка для сравнения')

    # TFS specific
    parser.add_argument('--username', type=str, required=False, help='Имя пользователя TFS')
    parser.add_argument('--password', type=str, required=False, help='Пароль TFS')
    parser.add_argument('--project', type=str, required=False, help='TFS Проект')
    parser.add_argument('--pat', type=str, required=False, help='PAT для TFS/Git clone')

    # GitHub specific
    parser.add_argument('--owner', type=str, required=False, help='Владелец репозитория GitHub')
    parser.add_argument('--github_token', type=str, required=False, help='GitHub Personal Access Token')

    args = parser.parse_args()

    if args.platform == 'tfs':
        if not all([args.username, args.password, args.project, args.pat]):
            parser.error("Для TFS требуются: --username, --password, --project, --pat")
    elif args.platform == 'github':
        if not all([args.owner, args.github_token]):
            parser.error("Для GitHub требуются: --owner, --github_token")

    if args.search_by_date.lower() in ('y', 'yes', 't', 'true', 'on', '1'):
        args.search_by_date = True
    elif args.search_by_date.lower() in ('n', 'no', 'f', 'false', 'off', '0'):
        args.search_by_date = False
    else:
        raise ValueError("Неверное значение для --search_by_date")

    return args

def filter_commits_by_date(commits, start_date, end_date):
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    filtered = []
    for commit in commits:
        commit_date = datetime.strptime(commit["author"]["date"], "%Y-%m-%dT%H:%M:%SZ").date()
        if start <= commit_date <= end:
            filtered.append(commit)
    return filtered

def filter_commits_by_branch(provider, commits):
    commit_ids = provider.get_commit_ids_from_branches()
    necessary_commits = [commit for commit in commits if commit["commitId"] in commit_ids]
    return necessary_commits

def filling_data(necessary_commits, provider, auth, headers):
    fields = ["Описание изменений", "ID коммита", "Автор изменений", "Дата", "Рабочий элемент", "Requirement", "Feature"]
    rows = []
    for commit in necessary_commits:
        if commit.get("workItems"):
            work_items = []
            requirements = []
            features = []

            for item in commit["workItems"]:
                parent_id, current_id, task_type, task_title, task_url = provider.get_work_item_info(item["id"], auth, headers)
                work_items.append(f"{task_url}!@#$%$#@!{task_type} {current_id}: {task_title}")

                if parent_id != "-":
                    parent_id, current_id, task_type, task_title, task_url = provider.get_work_item_info(parent_id, auth, headers)
                    requirements.append(f"{task_url}!@#$%$#@!{task_type} {current_id}: {task_title}")

                    if parent_id != "-":
                        parent_id, current_id, task_type, task_title, task_url = provider.get_work_item_info(parent_id, auth, headers)
                        features.append(f"{task_url}!@#$%$#@!{task_type} {current_id}: {task_title}")
                    else:
                        features.append("-")
                else:
                    requirements.append("-")
                    features.append("-")

            max_len = max(len(work_items), len(requirements), len(features))
            for i in range(max_len):
                row = [
                    commit["comment"] if i == 0 else "",
                    commit["commitId"] if i == 0 else "",
                    commit["author"]["name"] if i == 0 else "",
                    datetime.strptime(commit["author"]["date"], "%Y-%m-%dT%H:%M:%SZ").strftime("%d.%m.%Y") if i == 0 else "",
                    work_items[i] if i < len(work_items) else "-",
                    requirements[i] if i < len(requirements) else "-",
                    features[i] if i < len(features) else "-"
                ]
                rows.append(row)
        else:
            rows.append([
                commit["comment"], commit["commitId"], commit["author"]["name"],
                datetime.strptime(commit["author"]["date"], "%Y-%m-%dT%H:%M:%SZ").strftime("%d.%m.%Y"),
                "-", "-", "-"
            ])
    return fields, rows

def filling_data_in_xlsx(fields, rows, args):
    today = datetime.now().strftime("%d.%m.%Y")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = today
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 48
    ws.column_dimensions['F'].width = 43
    ws.column_dimensions['G'].width = 43

    for col_num, field in enumerate(fields, start=1):
        cell = ws.cell(row=2, column=col_num, value=field)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_offset = 3
    for row_num, row in enumerate(rows, start=row_offset):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num)

            if col_num in (5, 6, 7):
                if value != "-":
                    temp = value.split('!@#$%$#@!', 1)
                    cell.value = temp[-1]
                    cell.hyperlink = temp[0]
                    cell.style = "Hyperlink"
                else:
                    cell.value = value
            else:
                cell.value = value
            cell.alignment = Alignment(wrap_text=True)

    current_commit_start = row_offset
    for row_num in range(row_offset, len(rows) + row_offset):
        if not rows[row_num - row_offset][0]:
            continue
        if row_num > current_commit_start:
            for col in range(1, 5):
                ws.merge_cells(start_row=current_commit_start, end_row=row_num - 1, start_column=col, end_column=col)
        current_commit_start = row_num

    if current_commit_start < len(rows) + row_offset:
        for col in range(1, 5):
            ws.merge_cells(start_row=current_commit_start, end_row=len(rows) + row_offset - 1, start_column=col, end_column=col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    if args.search_by_date:
        ru_start_date = datetime.strptime(args.start_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        ru_end_date = datetime.strptime(args.end_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        cell = ws.cell(row=1, column=1, value=f"Дата создания отчета:{today}, Репозиторий:{args.repository}, Ветка: {args.branch}, Отчет за период: {ru_start_date}-{ru_end_date}")
    else:
        cell = ws.cell(row=1, column=1, value=f"Дата создания отчета:{today}, Репозиторий:{args.repository}, Отчет об изменениях между ветками: {args.previous_branch} и {args.branch}")
    wb.save(f"{today}_{args.repository}.xlsx")

def main():
    args = parse_arguments()
    if args.platform == 'tfs':
        provider = TFSProvider(args)
    elif args.platform == 'github':
        provider = GitHubProvider(args)
    else:
        raise ValueError("Неверная платформа")

    if args.search_by_date:
        if not args.start_date:
            args.start_date = "1970-01-01"  # Default to far past
        if not args.end_date or args.end_date == "-":
            args.end_date = datetime.now().strftime("%Y-%m-%d")

    api_url = provider.construct_api_url()
    auth, headers = provider.get_auth_headers()
    if isinstance(api_url, tuple):
        api_url, params = api_url
    else:
        params = {}
    commits = provider.fetch_commits(api_url, auth, headers)

    if args.search_by_date:
        necessary_commits = filter_commits_by_date(commits, args.start_date, args.end_date)
    else:
        necessary_commits = filter_commits_by_branch(provider, commits)

    if not necessary_commits:
        print("Нет коммитов по критериям.")
        sys.exit()

    fields, rows = filling_data(necessary_commits, provider, auth, headers)
    filling_data_in_xlsx(fields, rows, args)

if __name__ == "__main__":
    main()